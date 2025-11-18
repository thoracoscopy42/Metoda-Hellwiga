import pandas as pd
import itertools as it
from pathlib import Path
from openpyxl import load_workbook
import customtkinter as ctk
from tkinter import filedialog, messagebox



FILEPATH = Path(r"dummy_excel.xlsx")


"""

Hellwig Method Integral Capacity Calculation


"""
def check_y_column(df):

    # formatting fixes
    df.columns = [col.strip() for col in df.columns]
    df.columns = [col.replace(" ", "_") for col in df.columns]
    df.columns = [col.upper() for col in df.columns]
    df.index = df.columns # set index same as columns
    # check for Y
    if 'Y' not in df.columns:
        raise ValueError("Kolumna Y jest potrzebna do obliczenia pojemności integralnej.")
    
    return df
    
def separate_x_y(df):
    x_cols = [c for c in df.columns if c != 'Y']
    # R = df.drop(columns=['Y'])
    # R0 = df['Y']
    # R0 = R0 ** 2 # square the Y correlation values for easier formula implementation
    R = df.loc[x_cols, x_cols]
    R0 = df.loc[x_cols, 'Y'] ** 2

    return R, R0


def calculate_integral_capacity(R, R0):
    H_values = {} # dictionary to store integral capacities for each combination

    for size in range(1, len(R.columns) + 1): # all combination sizes
        for combo in it.combinations(R.columns, size): # all combinations of that size
            H_sum = 0 # sum of integral capacities for current combination
            R_subset = R.loc[list(combo), list(combo)] # current combination subset
            for j in combo:
                denom = sum(abs(R_subset.loc[j, i]) for i in combo)
                H_j = R0[j] / denom 
                H_sum += H_j
            H_values[combo] = H_sum

    return H_values

def output_results_to_excel(H_values, FILEPATH):
    output_df = pd.DataFrame(
        list(H_values.items()),
        columns=['Combination', 'Integral_Capacity']
    )

    sheet_name = "Integral_Capacity_Results"

    try:
        wb = load_workbook(FILEPATH)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]       
            wb.save(FILEPATH)
    except FileNotFoundError:
        pass

    with pd.ExcelWriter(FILEPATH, mode='a', engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name=sheet_name)

    print(f"Wyniki zostały zapisane do arkusza '{sheet_name}' w {FILEPATH}")

def browse_file():
    fname = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xls")],
        title="Wybierz plik z macierzą korelacji"
    )
    if fname:
        file_var.set(fname)
        status_var.set("Wybrano plik.")


"""
GUI

"""



def run_hellwig():
    path_str = file_var.get().strip()
    if not path_str:
        messagebox.showwarning("Uwaga", "Wybierz plik Excela.")
        return

    filepath = Path(path_str)
    if not filepath.exists():
        messagebox.showerror("Błąd", f"Plik {filepath} nie istnieje.")
        return

    try:
        status_var.set("Trwa obliczanie...")
        app.update_idletasks()

        # core
        df = pd.read_excel(filepath, sheet_name=0)
        df = check_y_column(df)
        R, R0 = separate_x_y(df)
        H_values = calculate_integral_capacity(R, R0)
        output_results_to_excel(H_values, filepath)

        # pokaż kilka top kombinacji w UI
        items_sorted = sorted(H_values.items(), key=lambda x: x[1], reverse=True)
        top_n = 10 if len(items_sorted) >= 10 else len(items_sorted)
        top_items = items_sorted[:top_n]

        result_box.configure(state="normal")
        result_box.delete("1.0", "end")
        result_box.insert("end", f"Plik: {filepath}\n")
        result_box.insert("end", f"Liczba kombinacji: {len(H_values)}\n")
        result_box.insert("end", f"TOP {top_n} kombinacji:\n\n")
        for combo, H_val in top_items:
            result_box.insert("end", f"{combo}  ->  {H_val:.6f}\n")
        result_box.configure(state="disabled")

        status_var.set("Gotowe. Wyniki zapisane w Excelu.")
        # messagebox.showinfo(
        #     "Sukces",
        #     f"Wyniki zapisano w arkuszu 'Integral_Capacity_Results' w pliku:\n{filepath}"
        # )

    except Exception as e:
        status_var.set("Błąd.")
        messagebox.showerror("Błąd", str(e))


if __name__ == "__main__":
    # wygląd UI
    ctk.set_appearance_mode("dark")         # "light", "dark", "system"
    ctk.set_default_color_theme("dark-blue")       # "blue", "green", "dark-blue"

    app = ctk.CTk()
    app.title("Metoda Hellwiga")
    app.geometry("900x600")

    file_var = ctk.StringVar()
    status_var = ctk.StringVar(value="Gotowy.")

    main_frame = ctk.CTkFrame(app, corner_radius=15)
    main_frame.pack(fill="both", expand=True, padx=20, pady=20)

    title_label = ctk.CTkLabel(
        main_frame,
        text="Metoda Hellwiga – obliczanie pojemności informacyjnej",
        font=ctk.CTkFont(size=16, weight="bold")
    )
    title_label.grid(row=0, column=0, columnspan=2, pady=(10, 15), sticky="w")

    # wybór pliku
    file_label = ctk.CTkLabel(main_frame, text="Plik Excel z macierzą korelacji:")
    file_label.grid(row=1, column=0, sticky="w")

    file_entry = ctk.CTkEntry(main_frame, textvariable=file_var, width=420)
    file_entry.grid(row=2, column=0, pady=5, sticky="we")

    browse_button = ctk.CTkButton(main_frame, text="Przeglądaj...", command=browse_file)
    browse_button.grid(row=2, column=1, padx=(10, 0), pady=5, sticky="e")

    # przycisk start
    run_button = ctk.CTkButton(
        main_frame,
        text="Uruchom metodę Hellwiga",
        command=run_hellwig,
        height=36
    )
    run_button.grid(row=3, column=0, columnspan=2, pady=(15, 10), sticky="we")

    # pole na wyniki
    result_box = ctk.CTkTextbox(main_frame, height=180)
    result_box.grid(row=4, column=0, columnspan=2, pady=(10, 5), sticky="nsew")
    result_box.configure(state="disabled")

    # status
    status_label = ctk.CTkLabel(main_frame, textvariable=status_var, anchor="w")
    status_label.grid(row=5, column=0, columnspan=2, sticky="we", pady=(5, 0))

    # layout stretch
    main_frame.columnconfigure(0, weight=3)
    main_frame.columnconfigure(1, weight=1)
    main_frame.rowconfigure(4, weight=1)

    app.mainloop()