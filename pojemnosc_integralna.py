import pandas as pd
import itertools as it
from pathlib import Path
from openpyxl import load_workbook



FILEPATH = Path(r"dummy_excel.xlsx")

def check_y_column(df):

    # formatting fixes
    df.columns = [col.strip() for col in df.columns]
    df.columns = [col.replace(" ", "_") for col in df.columns]
    df.columns = [col.upper() for col in df.columns]
    df.index = df.columns # set index same as columns
    # check for Y
    if 'Y' not in df.columns:
        raise ValueError("Kolumna Y jest potrzebna do obliczenia pojemności integralnej.")
    
    return df
    
def separate_x_y(df):
    x_cols = [c for c in df.columns if c != 'Y']
    # R = df.drop(columns=['Y'])
    # R0 = df['Y']
    # R0 = R0 ** 2 # square the Y correlation values for easier formula implementation
    R = df.loc[x_cols, x_cols]
    R0 = df.loc[x_cols, 'Y'] ** 2

    return R, R0


def calculate_integral_capacity(R, R0):
    H_values = {} # dictionary to store integral capacities for each combination

    for size in range(1, len(R.columns) + 1): # all combination sizes
        for combo in it.combinations(R.columns, size): # all combinations of that size
            H_sum = 0 # sum of integral capacities for current combination
            R_subset = R.loc[list(combo), list(combo)] # current combination subset
            for j in combo:
                denom = sum(abs(R_subset.loc[j, i]) for i in combo)
                H_j = R0[j] / denom 
                H_sum += H_j
            H_values[combo] = H_sum

    return H_values

def output_results_to_excel(H_values):
    output_df = pd.DataFrame(
        list(H_values.items()),
        columns=['Combination', 'Integral_Capacity']
    )

    sheet_name = "Integral_Capacity_Results"

    
    try:
        wb = load_workbook(FILEPATH)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]       
            wb.save(FILEPATH)

    except FileNotFoundError:
        pass

    with pd.ExcelWriter(FILEPATH, mode='a', engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name=sheet_name)

    print(f"Wyniki zostały zapisane do arkusza '{sheet_name}' w {FILEPATH}")


if __name__ == "__main__":

    df = pd.read_excel(FILEPATH, sheet_name=0)
    check_y_column(df) 
    R, R0 = separate_x_y(df)
    H_values = calculate_integral_capacity(R, R0)
    output_results_to_excel(H_values)
    
    
    
